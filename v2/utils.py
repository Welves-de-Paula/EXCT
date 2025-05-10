# -*- coding: utf-8 -*-
import os
import shutil
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from copy import copy
import subprocess

def dividir_arquivo_excel(caminho_arquivo, max_linhas, manter_formatacao, clonar_cabecalho, queue):
    try:
        wb = load_workbook(caminho_arquivo)
        ws = wb.active

        if ws.max_row < 2:
            queue.put(("log", "O arquivo não contém dados suficientes para dividir."))
            return

        nome_arquivo = os.path.splitext(os.path.basename(caminho_arquivo))[0]
        diretorio_saida = os.path.join(os.path.dirname(caminho_arquivo), f"EXCT_{nome_arquivo}")

        if os.path.exists(diretorio_saida):
            shutil.rmtree(diretorio_saida)

        os.makedirs(diretorio_saida, exist_ok=True)

        total_linhas = ws.max_row - 1
        total_partes = (total_linhas + max_linhas - 1) // max_linhas

        for parte in range(total_partes):
            novo_wb = Workbook()
            novo_ws = novo_wb.active

            if clonar_cabecalho:
                for col in range(1, ws.max_column + 1):
                    cell_origem = ws.cell(row=1, column=col)
                    cell_destino = novo_ws.cell(row=1, column=col, value=cell_origem.value)
                    if manter_formatacao:
                        cell_destino.font = copy(cell_origem.font)
                        cell_destino.fill = copy(cell_origem.fill)
                        cell_destino.border = copy(cell_origem.border)
                        cell_destino.alignment = copy(cell_origem.alignment)
                        cell_destino.number_format = copy(cell_origem.number_format)

            inicio = parte * max_linhas + 2
            fim = min((parte + 1) * max_linhas + 2, total_linhas + 2)

            for linha in range(inicio, fim):
                for col in range(1, ws.max_column + 1):
                    cell_origem = ws.cell(row=linha, column=col)
                    cell_destino = novo_ws.cell(row=linha - inicio + 2, column=col, value=cell_origem.value)
                    if manter_formatacao:
                        cell_destino.font = copy(cell_origem.font)
                        cell_destino.fill = copy(cell_origem.fill)
                        cell_destino.border = copy(cell_origem.border)
                        cell_destino.alignment = copy(cell_origem.alignment)
                        cell_destino.number_format = copy(cell_origem.number_format)

            for col in range(1, ws.max_column + 1):
                col_letter = get_column_letter(col)
                if ws.column_dimensions[col_letter].width:
                    novo_ws.column_dimensions[col_letter].width = ws.column_dimensions[col_letter].width

            caminho_parte = os.path.join(diretorio_saida, f"{nome_arquivo}_parte_{parte + 1}.xlsx")
            novo_wb.save(caminho_parte)
            queue.put(("log", f"Arquivo gerado: {caminho_parte}"))

            progresso_percentual = ((parte + 1) / total_partes) * 100
            queue.put(("progresso", progresso_percentual))

        queue.put(("log", "Tarefa concluída com sucesso."))
        queue.put(("final", (total_linhas, total_partes, diretorio_saida)))
    except Exception as e:
        queue.put(("log", f"Erro: {e}"))
    finally:
        queue.put(("progresso", 100))

def abrir_diretorio(diretorio_saida):
    subprocess.Popen(["explorer", os.path.realpath(diretorio_saida)])
