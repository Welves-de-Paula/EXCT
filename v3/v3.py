import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from tkinter.ttk import Progressbar
import pandas as pd
import logging
import os
import re
from datetime import datetime

# Configuração inicial do logger
logging.basicConfig(
    level=logging.INFO,
    format='[%(asctime)s] %(message)s',
    datefmt='%d/%m/%Y %H:%M',
)

# ==========================
# Funções de Leitura de Arquivos
# ==========================
def carregar_arquivo():
    """Abre um diálogo para selecionar um arquivo Excel e carrega os dados em um DataFrame."""
    filepath = filedialog.askopenfilename(
        filetypes=[("Arquivos Excel", "*.xlsx")],
        title="Selecione um arquivo Excel"
    )
    if filepath:
        try:
            global df
            df = pd.read_excel(filepath, engine='openpyxl')  # Usa openpyxl para leitura
            logging.info(f"Arquivo {filepath} carregado com sucesso.")
            messagebox.showinfo("Sucesso", "Arquivo carregado com sucesso!")
        except Exception as e:
            logging.error(f"Erro ao carregar arquivo: {e}")
            messagebox.showerror("Erro", f"Erro ao carregar arquivo: {e}")

def salvar_arquivo():
    """Salva o DataFrame carregado em um novo arquivo Excel."""
    if 'df' not in globals():
        messagebox.showwarning("Aviso", "Nenhum arquivo carregado para salvar.")
        return
    filepath = filedialog.asksaveasfilename(
        defaultextension=".xlsx",
        filetypes=[("Arquivos Excel", "*.xlsx")],
        title="Salvar arquivo como"
    )
    if filepath:
        try:
            df.to_excel(filepath, index=False, engine='openpyxl')
            logging.info(f"Arquivo salvo com sucesso em {filepath}.")
            messagebox.showinfo("Sucesso", "Arquivo salvo com sucesso!")
        except Exception as e:
            logging.error(f"Erro ao salvar arquivo: {e}")
            messagebox.showerror("Erro", f"Erro ao salvar arquivo: {e}")

# ==========================
# Funções de Manipulação de Dados
# ==========================
def limpar():
    """Remove o DataFrame carregado da memória."""
    global df
    if 'df' in globals():
        del df
        logging.info("Dados limpos da memória.")
        messagebox.showinfo("Sucesso", "Dados limpos da memória.")
    else:
        logging.warning("Nenhum dado carregado para limpar.")
        messagebox.showwarning("Aviso", "Nenhum dado carregado para limpar.")

def validar():
    """Valida os dados do DataFrame carregado conforme regras definidas e gera arquivo pós-validação."""
    if 'df' not in globals():
        messagebox.showwarning("Aviso", "Nenhum arquivo carregado para validar.")
        return
    try:
        global df
        erros = []
        df_validado = df.copy()

        for index, row in df.iterrows():
            # Validação para tabela de clientes
            if 'NOME' in df.columns:
                if not isinstance(row.get('NOME'), str) or not (3 <= len(row['NOME']) <= 50):
                    erros.append((index + 1, "Nome inválido"))
                if 'CPF' in df.columns and pd.notna(row.get('CPF')):
                    if not re.fullmatch(r'\d{11}', str(row['CPF'])):
                        erros.append((index + 1, "CPF inválido"))
                if 'CNPJ' in df.columns and pd.notna(row.get('CNPJ')):
                    if not re.fullmatch(r'\d{14}', str(row['CNPJ'])):
                        erros.append((index + 1, "CNPJ inválido"))
                if 'EMAIL' in df.columns and pd.notna(row.get('EMAIL')):
                    if not re.fullmatch(r"[^@]+@[^@]+\.[^@]+", row['EMAIL']):
                        erros.append((index + 1, "Email inválido"))
                if 'ATIVO' in df.columns:
                    if str(row.get('ATIVO')).upper() not in ['S', 'N', 'SIM', 'NÃO', '1', '0']:
                        erros.append((index + 1, "Valor inválido para ATIVO"))
                if 'DATA DE NASCIMENTO' in df.columns and pd.notna(row.get('DATA DE NASCIMENTO')):
                    try:
                        datetime.strptime(row['DATA DE NASCIMENTO'], '%d/%m/%Y')
                    except ValueError:
                        erros.append((index + 1, "Data de nascimento inválida"))

            # Validação para tabela de produtos
            if 'NOME' in df.columns and 'CÓDIGO DE BARRAS' in df.columns:
                if not isinstance(row.get('NOME'), str) or not (3 <= len(row['NOME']) <= 50):
                    erros.append((index + 1, "Nome do produto inválido"))
                if 'CÓDIGO DE BARRAS' in df.columns and pd.notna(row.get('CÓDIGO DE BARRAS')):
                    if not re.fullmatch(r'\d{13}', str(row['CÓDIGO DE BARRAS'])):
                        erros.append((index + 1, "Código de barras inválido"))
                if 'PREÇO' in df.columns and pd.notna(row.get('PREÇO')):
                    if not (0 <= row['PREÇO'] <= 999999.99):
                        erros.append((index + 1, "Preço fora do intervalo permitido"))
                if 'UNIDADE DE VENDA' in df.columns:
                    if row.get('UNIDADE DE VENDA') not in ['UNID', 'KG', 'M', 'LITRO']:
                        erros.append((index + 1, "Unidade de venda inválida"))

        # Adicionar coluna de erros no DataFrame
        if erros:
            df_validado['Motivo da Invalidação'] = ""
            for linha, motivo in erros:
                df_validado.at[linha - 1, 'Motivo da Invalidação'] = motivo

            # Aplicar formatação para linhas inválidas
            from openpyxl import load_workbook
            from openpyxl.styles import PatternFill

            arquivo_validado = "validado.xlsx"
            df_validado.to_excel(arquivo_validado, index=False, engine='openpyxl')

            wb = load_workbook(arquivo_validado)
            ws = wb.active
            fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")

            for linha, _ in erros:
                for cell in ws[linha + 1]:  # +1 para ajustar ao cabeçalho
                    cell.fill = fill

            wb.save(arquivo_validado)
            logging.error(f"Validação concluída com {len(erros)} erro(s). Arquivo salvo como {arquivo_validado}.")
            messagebox.showerror("Erro", "Validação falhou. Verifique o arquivo validado e os logs.")
        else:
            logging.info("Arquivo validado com sucesso! Nenhum erro encontrado.")
            messagebox.showinfo("Sucesso", "Arquivo validado com sucesso!")

    except Exception as e:
        logging.error(f"Erro durante validação: {e}")
        messagebox.showerror("Erro", f"Erro durante validação: {e}")

def comparar():
    """Compara o DataFrame carregado com outro arquivo Excel."""
    if 'df' not in globals():
        messagebox.showwarning("Aviso", "Nenhum arquivo carregado para comparar.")
        return
    filepath = filedialog.askopenfilename(
        filetypes=[("Arquivos Excel", "*.xlsx")],
        title="Selecione um segundo arquivo Excel"
    )
    if filepath:
        try:
            df2 = pd.read_excel(filepath, engine='openpyxl')
            diferencas = pd.concat([df, df2]).drop_duplicates(keep=False)
            if not diferencas.empty:
                logging.info(f"Diferenças encontradas entre os arquivos. Total de diferenças: {len(diferencas)}.")
                messagebox.showinfo("Comparação", "Diferenças encontradas. Verifique os logs.")
            else:
                logging.info("Os arquivos são idênticos.")
                messagebox.showinfo("Comparação", "Os arquivos são idênticos.")
        except Exception as e:
            logging.error(f"Erro ao comparar arquivos: {e}")
            messagebox.showerror("Erro", f"Erro ao comparar arquivos: {e}")

def dividir():
    """Divide o DataFrame carregado em múltiplos arquivos menores."""
    if 'df' not in globals():
        messagebox.showwarning("Aviso", "Nenhum arquivo carregado para dividir.")
        return
    try:
        # Configuração inicial
        linhas_por_arquivo = tk.simpledialog.askinteger(
            "Dividir Arquivo", "Digite a quantidade de linhas por arquivo (padrão: 1000):", initialvalue=1000
        )
        if not linhas_por_arquivo or linhas_por_arquivo <= 0:
            messagebox.showerror("Erro", "Quantidade de linhas inválida.")
            return

        manter_formatacao = messagebox.askyesno("Dividir Arquivo", "Manter formatação?")
        clonar_cabecalho = messagebox.askyesno("Dividir Arquivo", "Clonar cabeçalho?")

        # Criar subpasta para salvar os arquivos divididos
        subpasta = f"{os.path.splitext(os.path.basename(filepath))[0]}_div"
        os.makedirs(subpasta, existindo=True)

        # Divisão do DataFrame
        for i, chunk in enumerate(range(0, len(df), linhas_por_arquivo)):
            arquivo_saida = os.path.join(subpasta, f"parte_{i + 1}.xlsx")
            df_chunk = df.iloc[chunk:chunk + linhas_por_arquivo]

            if clonar_cabecalho:
                df_chunk.to_excel(arquivo_saida, index=False, engine='openpyxl')
            else:
                df_chunk.to_excel(arquivo_saida, index=False, header=False, engine='openpyxl')

            logging.info(f"Parte {i + 1} salva em {arquivo_saida}.")

        logging.info(f"Divisão concluída. Total de partes geradas: {i + 1}.")
        messagebox.showinfo("Sucesso", f"Arquivo dividido com sucesso! Arquivos salvos na pasta '{subpasta}'.")
    except Exception as e:
        logging.error(f"Erro ao dividir arquivo: {e}")
        messagebox.showerror("Erro", f"Erro ao dividir arquivo: {e}")

# ==========================
# Funções de Controle da Interface
# ==========================
def sair():
    """Fecha a aplicação."""
    logging.info("Aplicação encerrada pelo usuário.")
    root.destroy()

# ==========================
# Configuração da Interface Gráfica
# ==========================
root = tk.Tk()
root.title("Assistente de Importação de Arquivos Excel")
root.geometry("800x600")
root.resizable(True, True)

# Adicionar menu suspenso
menu_bar = tk.Menu(root)
menu_ajuda = tk.Menu(menu_bar, tearoff=0)
menu_ajuda.add_command(label="Sobre", command=lambda: messagebox.showinfo("Sobre", "Assistente de Importação v1.0"))
menu_bar.add_cascade(label="Ajuda", menu=menu_ajuda)
root.config(menu=menu_bar)

# Estilo (opcional: ttkbootstrap pode ser adicionado aqui)
frame = ttk.Frame(root, padding="10")
frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))

# Adicionar barra de progresso
progress = Progressbar(root, orient=tk.HORIZONTAL, length=400, mode='determinate')
progress.grid(row=2, column=0, pady=10)

# Função para atualizar a barra de progresso
def atualizar_progresso(valor):
    progress['value'] = valor
    root.update_idletasks()

# Botões com ícones e tooltips
def adicionar_tooltip(widget, texto):
    tooltip = tk.Toplevel(widget)
    tooltip.withdraw()
    tooltip.overrideredirect(True)
    tooltip_label = tk.Label(tooltip, text=texto, background="yellow", relief="solid", borderwidth=1, font=("Arial", 10))
    tooltip_label.pack()

    def mostrar_tooltip(event):
        tooltip.geometry(f"+{event.x_root + 10}+{event.y_root + 10}")
        tooltip.deiconify()

    def esconder_tooltip(event):
        tooltip.withdraw()

    widget.bind("<Enter>", mostrar_tooltip)
    widget.bind("<Leave>", esconder_tooltip)

btn_carregar = ttk.Button(frame, text="Carregar Arquivo", command=carregar_arquivo)
btn_carregar.grid(row=0, column=0, padx=5, pady=5)
adicionar_tooltip(btn_carregar, "Carregar um arquivo Excel (.xlsx)")

btn_salvar = ttk.Button(frame, text="Salvar", command=salvar_arquivo)
btn_salvar.grid(row=0, column=1, padx=5, pady=5)
adicionar_tooltip(btn_salvar, "Salvar o arquivo carregado")

btn_dividir = ttk.Button(frame, text="Dividir", command=dividir)
btn_dividir.grid(row=0, column=2, padx=5, pady=5)
adicionar_tooltip(btn_dividir, "Dividir o arquivo em partes menores")

btn_sair = ttk.Button(frame, text="Sair", command=sair)
btn_sair.grid(row=0, column=3, padx=5, pady=5)
adicionar_tooltip(btn_sair, "Fechar a aplicação")

btn_validar = ttk.Button(frame, text="Validar", command=validar)
btn_validar.grid(row=1, column=0, padx=5, pady=5)
adicionar_tooltip(btn_validar, "Validar os dados do arquivo carregado")

btn_comparar = ttk.Button(frame, text="Comparar", command=comparar)
btn_comparar.grid(row=1, column=1, padx=5, pady=5)
adicionar_tooltip(btn_comparar, "Comparar o arquivo carregado com outro arquivo Excel")

btn_limpar = ttk.Button(frame, text="Limpar", command=limpar)
btn_limpar.grid(row=1, column=2, padx=5, pady=5)
adicionar_tooltip(btn_limpar, "Limpar os dados carregados da memória")

# ==========================
# Configuração da Seção de Logs
# ==========================
log_frame = ttk.LabelFrame(root, text="Logs", padding="10")
log_frame.grid(row=1, column=0, sticky=(tk.W, tk.E, tk.N, tk.S), padx=10, pady=10)

log_text = tk.Text(log_frame, height=15, state="disabled", wrap="word", background="#f0f0f0")
log_text.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))

# Redirecionar logs para a interface
class TextHandler(logging.Handler):
    """Classe para redirecionar logs para a interface gráfica."""
    def emit(self, record):
        msg = self.format(record)
        log_text.configure(state="normal")
        log_text.insert(tk.END, msg + "\n")
        log_text.configure(state="disabled")
        log_text.see(tk.END)

handler = TextHandler()
logging.getLogger().addHandler(handler)

# ==========================
# Inicialização da Interface
# ==========================
root.mainloop()