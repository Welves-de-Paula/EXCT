import os
import re
from openpyxl import load_workbook


def normalize_header(header):
    """
    Normaliza o nome do header para snake_case.
    """
    header = header.strip().lower()
    header = re.sub(r"[áàãâä]", "a", header)
    header = re.sub(r"[éèêë]", "e", header)
    header = re.sub(r"[íìîï]", "i", header)
    header = re.sub(r"[óòõôö]", "o", header)
    header = re.sub(r"[úùûü]", "u", header)
    header = re.sub(r"[ç]", "c", header)
    header = re.sub(r"[^a-z0-9]+", "_", header)
    header = re.sub(r"_+", "_", header)
    header = header.strip("_")
    return header


def get_cell_style(cell):
    """
    Retorna um dicionário com informações de estilo da célula.
    """
    style = {}
    if cell.font:
        style['font'] = {
            'name': cell.font.name,
            'size': cell.font.size,
            'bold': cell.font.bold,
            'italic': cell.font.italic,
            'color': str(cell.font.color.rgb) if cell.font.color else None
        }
    if cell.fill:
        style['fill'] = {
            'fgColor': str(cell.fill.fgColor.rgb) if cell.fill.fgColor else None
        }
    if cell.border:
        style['border'] = str(cell.border)  # pode ser detalhado se necessário
    if cell.alignment:
        style['alignment'] = {
            'horizontal': cell.alignment.horizontal,
            'vertical': cell.alignment.vertical
        }
    return style


def read_excel(filepath):
    """
    Lê um arquivo Excel (.xlsx) e retorna um dicionário com headers, rows e style.
    Ignora linhas completamente vazias.
    """
    ext = os.path.splitext(filepath)[1].lower()
    if ext != ".xlsx":
        raise ValueError(
            "Arquivo não suportado. Por favor, utilize apenas arquivos .xlsx.")
    wb = load_workbook(filepath, read_only=False, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=False))
    if not rows:
        return {'headers': [], 'rows': [], 'style': {}}
    # Headers
    headers = [
        str(cell.value) if cell.value is not None else "" for cell in rows[0]]
    normalized_headers = [normalize_header(h) for h in headers]

    def idx_to_excel_col(idx):
        col = ''
        while idx >= 0:
            col = chr(idx % 26 + 65) + col
            idx = idx // 26 - 1
        return col
    headers_dict = [
        {
            'key': nh,
            'label': h,
            'column': idx_to_excel_col(idx)
        }
        for idx, (h, nh) in enumerate(zip(headers, normalized_headers))
    ]    # Rows (ignora linhas completamente vazias)
    data_rows = []
    for r_idx, row in enumerate(rows[1:]):
        row_dict = {}
        is_empty = True
        for c_idx, cell in enumerate(row):
            key = normalized_headers[c_idx] if c_idx < len(
                normalized_headers) else f'col_{c_idx+1}'
            value = str(cell.value) if cell.value is not None else ""
            label = headers[c_idx] if c_idx < len(headers) else key
            cell_ref = f"{chr(65 + c_idx)}{r_idx+2}"

            style = get_cell_style(cell)
            if value.strip() != "":
                is_empty = False
            row_dict[key] = {
                'key': key,
                'value': value,
                'label': label,
                'cell': cell_ref,
                'column': idx_to_excel_col(c_idx),
                'row': r_idx + 2,  # Linha do Excel (1-indexed)
                'style': style
            }

        if not is_empty:
            # Adiciona referência da linha do Excel
            row_dict['row_ref'] = r_idx + 2
            data_rows.append(row_dict)

    return {
        'headers': headers_dict,
        'rows': data_rows,
    }
