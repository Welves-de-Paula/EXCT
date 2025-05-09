import os
import threading
import time
from tkinter import Tk, filedialog, Toplevel, Label, Button, Text, Scrollbar, END, StringVar, messagebox
from tkinter.ttk import Progressbar
import subprocess
from copy import copy
import shutil

try:
    import pandas as pd
except ModuleNotFoundError:
    print("O módulo 'pandas' não está instalado. Por favor, instale-o usando o comando:")
    print("pip install pandas openpyxl")
    exit()

from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter

import queue

class App:
    def __init__(self, root):
        self.root = root
        self.root.title("Divisor de Arquivos Excel")
        self.root.geometry("600x400")

        self.running = False
        self.queue = queue.Queue()

        self.select_button = Button(root, text="Selecionar Arquivo Excel", command=self.selecionar_arquivo)
        self.select_button.pack(pady=10)

        self.progress_label = Label(root, text="Progresso: 0%")
        self.progress_label.pack(pady=5)
        self.progress = Progressbar(root, mode="determinate", maximum=100)
        self.progress.pack(pady=10, fill="x")

        self.log_area = Text(root, height=15, state="disabled", wrap="word")
        self.log_area.pack(pady=10, fill="both", expand=True)
        scrollbar = Scrollbar(self.log_area, command=self.log_area.yview)
        self.log_area.configure(yscrollcommand=scrollbar.set)

        self.atualizar_interface()

    def selecionar_arquivo(self):
        caminho_arquivo = filedialog.askopenfilename(
            title="Selecione o arquivo Excel",
            filetypes=[("Arquivos Excel", "*.xlsx")]
        )
        if caminho_arquivo:
            self.queue.put(("log", f"Arquivo selecionado: {caminho_arquivo}"))
            threading.Thread(target=self.dividir_arquivo_excel, args=(caminho_arquivo,)).start()
        else:
            self.queue.put(("log", "Nenhum arquivo selecionado."))

    def dividir_arquivo_excel(self, caminho_arquivo):
        try:
            wb = load_workbook(caminho_arquivo)
            ws = wb.active

            if ws.max_row < 2:
                self.queue.put(("log", "O arquivo não contém dados suficientes para dividir."))
                return

            nome_arquivo = os.path.splitext(os.path.basename(caminho_arquivo))[0]
            diretorio_saida = os.path.join(os.path.dirname(caminho_arquivo), f"EXCT_{nome_arquivo}")
            
            # Verificar se o diretório já existe e deletá-lo
            if os.path.exists(diretorio_saida):
                shutil.rmtree(diretorio_saida)
            
            os.makedirs(diretorio_saida, exist_ok=True)

            max_linhas = 1000
            total_linhas = ws.max_row - 1
            total_partes = (total_linhas + max_linhas - 1) // max_linhas

            for parte in range(total_partes):
                novo_wb = Workbook()
                novo_ws = novo_wb.active

                for col in range(1, ws.max_column + 1):
                    cell_origem = ws.cell(row=1, column=col)
                    cell_destino = novo_ws.cell(row=1, column=col, value=cell_origem.value)
                    cell_destino.font = copy(cell_origem.font)
                    cell_destino.fill = copy(cell_origem.fill)
                    cell_destino.border = copy(cell_origem.border)
                    cell_destino.alignment = copy(cell_origem.alignment)
                    cell_destino.number_format = copy(cell_origem.number_format)

                inicio = parte * max_linhas + 2
                fim = min((parte + 1) * max_linhas + 2, total_linhas + 2)

                for linha in range(inicio, fim):
                    for col in range(1, ws.max_column + 1):
                        cell_origem = ws.cell(row=linha, column=col)
                        cell_destino = novo_ws.cell(row=linha - inicio + 2, column=col, value=cell_origem.value)
                        cell_destino.font = copy(cell_origem.font)
                        cell_destino.fill = copy(cell_origem.fill)
                        cell_destino.border = copy(cell_origem.border)
                        cell_destino.alignment = copy(cell_origem.alignment)
                        cell_destino.number_format = copy(cell_origem.number_format)

                for col in range(1, ws.max_column + 1):
                    col_letter = get_column_letter(col)
                    if ws.column_dimensions[col_letter].width:
                        novo_ws.column_dimensions[col_letter].width = ws.column_dimensions[col_letter].width

                caminho_parte = os.path.join(diretorio_saida, f"{nome_arquivo}_parte_{parte + 1}.xlsx")
                novo_wb.save(caminho_parte)
                self.queue.put(("log", f"Arquivo gerado: {caminho_parte}"))

                progresso_percentual = ((parte + 1) / total_partes) * 100
                self.queue.put(("progresso", progresso_percentual))

            self.queue.put(("log", "Tarefa concluída com sucesso."))
            self.queue.put(("final", (total_linhas, total_partes, diretorio_saida)))
        except Exception as e:
            self.queue.put(("log", f"Erro: {e}"))
        finally:
            self.queue.put(("progresso", 100))

    def atualizar_interface(self):
        try:
            while not self.queue.empty():
                tipo, valor = self.queue.get(False)
                if tipo == "log":
                    self.log(valor)
                elif tipo == "progresso":
                    self.progress["value"] = valor
                    self.progress_label.config(text=f"Progresso: {int(valor)}%")
                elif tipo == "final":
                    self.exibir_dialogo_final(*valor)
        except queue.Empty:
            pass
        self.root.after(100, self.atualizar_interface)

    def exibir_dialogo_final(self, total_itens, total_tabelas, diretorio_saida):
        def abrir_diretorio():
            subprocess.Popen(["explorer", os.path.realpath(diretorio_saida)])

        messagebox.showinfo(
            "Processamento Concluído",
            f"Quantidade total de linhas escaneadas: {total_itens}\n"
            f"Quantidade de tabelas geradas: {total_tabelas}\n\n"
            f"Os arquivos foram salvos no diretório:\n{diretorio_saida}"
        )
        abrir_diretorio()

    def log(self, mensagem):
        self.log_area.config(state="normal")
        self.log_area.insert(END, mensagem + "\n")
        self.log_area.see(END)
        self.log_area.config(state="disabled")

if __name__ == "__main__":
    root = Tk()
    app = App(root)
    root.mainloop()
